from __future__ import annotations
import easyocr
import os
import re
import sqlite3
import uuid
from pathlib import Path
from typing import Dict, List, Optional

from flask import (
    Flask,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from mrz import extract_passport_info
from vin import extract_vin_info
from export_declaration import extract_export_declaration_info
from car365 import fetch_car365_vehicle_info

BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
STATIC_DIR = BASE_DIR / "static"
STAMPS_DIR = STATIC_DIR / "stamps"
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
TEMPLATE_PATH = BASE_DIR / "packing_list_template.xlsx"
DATABASE_PATH = INSTANCE_DIR / "app.db"

INSTANCE_DIR.mkdir(exist_ok=True)
STATIC_DIR.mkdir(exist_ok=True)
STAMPS_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

ALLOWED_STAMP_EXTENSIONS = {".png", ".jpg", ".jpeg"}

COUNTRY_OPTIONS = [
    "Kyrgyzstan",
    "Kazakhstan",
    "Uzbekistan",
    "Tajikistan",
    "Russia",
    "Mongolia",
    "Other",
]

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key")
app.config["MAX_CONTENT_LENGTH"] = 128 * 1024 * 1024


# ---------------- DB ----------------

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = get_db()
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS company_profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL UNIQUE,
            company_name TEXT DEFAULT '',
            company_address TEXT DEFAULT '',
            company_phone TEXT DEFAULT '',
            stamp_image_path TEXT DEFAULT '',
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        """
    )
    db.commit()


@app.before_request
def before_request():
    init_db()


# ---------------- Auth helpers ----------------

def current_user() -> Optional[sqlite3.Row]:
    user_id = session.get("user_id")
    if not user_id:
        return None
    db = get_db()
    return db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()


def login_required():
    if not current_user():
        flash("Please log in first.")
        return False
    return True


def get_company_profile(user_id: int) -> Optional[sqlite3.Row]:
    db = get_db()
    return db.execute(
        "SELECT * FROM company_profiles WHERE user_id = ?",
        (user_id,),
    ).fetchone()


def upsert_company_profile(
    user_id: int,
    company_name: str,
    company_address: str,
    company_phone: str,
    stamp_image_path: str,
):
    db = get_db()
    existing = get_company_profile(user_id)

    if existing:
        db.execute(
            """
            UPDATE company_profiles
            SET company_name = ?,
                company_address = ?,
                company_phone = ?,
                stamp_image_path = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE user_id = ?
            """,
            (company_name, company_address, company_phone, stamp_image_path, user_id),
        )
    else:
        db.execute(
            """
            INSERT INTO company_profiles (
                user_id, company_name, company_address, company_phone, stamp_image_path
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (user_id, company_name, company_address, company_phone, stamp_image_path),
        )

    db.commit()


# ---------------- General helpers ----------------

def safe_filename_part(value: str) -> str:
    value = (value or "").strip()
    allowed = "".join(ch for ch in value if ch.isalnum() or ch in ("-", "_"))
    return allowed or "packing_list"


def detect_file_role(filename: str) -> str | None:
    name = filename.lower()

    if name.endswith(".pdf"):
        if "export" in name or "declaration" in name or "exdec" in name:
            return "export"
        return "export"

    if "passport" in name or "mrz" in name or "pass" in name:
        return "passport"

    if "vin" in name or "sticker" in name or "label" in name:
        return "vin"

    return None


# ---------------- Pages ----------------

@app.route("/")
def home():
    user = current_user()
    profile = None
    if user:
        profile = get_company_profile(user["id"])
    return render_template("home.html", user=user, profile=profile)


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        confirm_password = request.form.get("confirm_password") or ""

        if not email:
            flash("Please enter email.")
            return redirect(url_for("register"))

        if not password:
            flash("Please enter password.")
            return redirect(url_for("register"))

        if password != confirm_password:
            flash("Passwords do not match.")
            return redirect(url_for("register"))

        db = get_db()
        existing = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
        if existing:
            flash("This email is already registered.")
            return redirect(url_for("register"))

        password_hash = generate_password_hash(password)
        db.execute(
            "INSERT INTO users (email, password_hash) VALUES (?, ?)",
            (email, password_hash),
        )
        db.commit()

        user = db.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        session["user_id"] = user["id"]

        flash("Registration completed.")
        return redirect(url_for("my_page"))

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""

        db = get_db()
        user = db.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()

        if user is None or not check_password_hash(user["password_hash"], password):
            flash("Invalid email or password.")
            return redirect(url_for("login"))

        session["user_id"] = user["id"]
        flash("Logged in successfully.")
        return redirect(url_for("home"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out.")
    return redirect(url_for("home"))


@app.route("/my-page", methods=["GET", "POST"])
def my_page():
    if not login_required():
        return redirect(url_for("login"))

    user = current_user()
    profile = get_company_profile(user["id"])

    if request.method == "POST":
        company_name = (request.form.get("company_name") or "").strip()
        company_address = (request.form.get("company_address") or "").strip()
        company_phone = (request.form.get("company_phone") or "").strip()

        stamp_image_path = ""
        if profile and profile["stamp_image_path"]:
            stamp_image_path = profile["stamp_image_path"]

        stamp_file = request.files.get("stamp_image")
        if stamp_file and stamp_file.filename:
            ext = Path(stamp_file.filename).suffix.lower()
            if ext not in ALLOWED_STAMP_EXTENSIONS:
                flash("Stamp image must be PNG or JPG.")
                return redirect(url_for("my_page"))

            stamp_filename = f"user_{user['id']}_{uuid.uuid4().hex}{ext}"
            saved_path = STAMPS_DIR / secure_filename(stamp_filename)
            stamp_file.save(saved_path)
            stamp_image_path = str(saved_path)

        upsert_company_profile(
            user_id=user["id"],
            company_name=company_name,
            company_address=company_address,
            company_phone=company_phone,
            stamp_image_path=stamp_image_path,
        )

        flash("Company profile saved.")
        return redirect(url_for("my_page"))

    profile = get_company_profile(user["id"])
    return render_template("my_page.html", user=user, profile=profile)


@app.route("/generator", methods=["GET", "POST"])
def generator():
    if not login_required():
        return redirect(url_for("login"))

    user = current_user()
    profile = get_company_profile(user["id"])

    if not profile:
        flash("Please fill company profile first.")
        return redirect(url_for("my_page"))

    if request.method == "GET":
        return render_template("generator.html", country_options=COUNTRY_OPTIONS, user=user)

    container_number = (request.form.get("container_number") or "").strip()
    seal_number = (request.form.get("seal_number") or "").strip()
    page_date = (request.form.get("page_date") or "").strip()
    country = (request.form.get("country") or "").strip()
    other_country = (request.form.get("other_country") or "").strip()
    city = (request.form.get("city") or "").strip()

    if country == "Other":
        country = other_country.strip()

    try:
        vehicle_count = int(request.form.get("vehicle_count", "0"))
    except ValueError:
        flash("Invalid number of vehicles.")
        return redirect(url_for("generator"))

    if not container_number:
        flash("Please enter container number.")
        return redirect(url_for("generator"))
    if not seal_number:
        flash("Please enter seal number.")
        return redirect(url_for("generator"))
    if not page_date:
        flash("Please select date.")
        return redirect(url_for("generator"))
    if not country:
        flash("Please select country.")
        return redirect(url_for("generator"))
    if not city:
        flash("Please enter city.")
        return redirect(url_for("generator"))
    if vehicle_count < 1 or vehicle_count > 10:
        flash("Vehicle count must be between 1 and 10.")
        return redirect(url_for("generator"))
    if not TEMPLATE_PATH.exists():
        flash("Packing list template not found.")
        return redirect(url_for("generator"))

    request_id = uuid.uuid4().hex
    destination = f"{country}, {city}"
    vehicle_rows: List[dict] = []

    for i in range(1, vehicle_count + 1):
        files = request.files.getlist(f"vehicle_files_{i}")
        valid_files = [f for f in files if f and f.filename]
        manual_vin = (
            request.form.get(f"vin_string_{i}") or ""
        ).strip()
        if not valid_files:
            flash(f"No files uploaded for vehicle #{i}.")
            return redirect(url_for("generator"))

        vehicle_dir = UPLOAD_DIR / request_id / f"vehicle_{i}"
        vehicle_dir.mkdir(parents=True, exist_ok=True)

        grouped: Dict[str, Path | None] = {
            "passport": None,
            "vin": None,
            "export": None,
        }

        for file in valid_files:
            safe_name = secure_filename(file.filename)
            role = detect_file_role(safe_name)

            if role is None:
                flash(f"File type not recognized for vehicle #{i}: {file.filename}")
                return redirect(url_for("generator"))

            file_path = vehicle_dir / safe_name
            file.save(file_path)
            grouped[role] = file_path

        passport_path = grouped["passport"]
        vin_path = grouped["vin"]
        export_path = grouped["export"]

        if passport_path is None:
            flash(f"Passport file missing for vehicle #{i}.")
            return redirect(url_for("generator"))
        if vin_path is None:
            flash(f"VIN sticker file missing for vehicle #{i}.")
            return redirect(url_for("generator"))
        if export_path is None:
            flash(f"Export declaration file missing for vehicle #{i}.")
            return redirect(url_for("generator"))

        passport_info = extract_passport_info(passport_path)
        vin_info = extract_vin_info(vin_path)
        export_info = extract_export_declaration_info(export_path)

        vin_string = (vin_info.get("vin_string") or "").strip()

        car365_info = {
            "returned_chassis_no": "",
            "vehicle_trademark": "",
            "vehicle_first_registration": "",
            "vehicle_fuel": "",
            "vehicle_displacement": "",
            "status": "SKIPPED",
            "error": "VIN is empty",
        }

        if vin_string:
            car365_info = fetch_car365_vehicle_info(vin_string)

        surname = (passport_info.get("surname") or "").strip()
        given_name = (passport_info.get("given_name") or "").strip()
        full_name = " ".join(part for part in [surname, given_name] if part).strip()

        vehicle_rows.append(
            {
                "vehicle_no": i,
                "kg_number": f"KG-{i:02d}",
                "page_date": page_date,
                "container_number": container_number,
                "seal_number": seal_number,
                "destination": destination,
                "full_name": full_name,
                "document_id": (passport_info.get("document_id") or "").strip(),
                "vin_string": manual_vin,
                "hs_code": (export_info.get("hs_code") or "").strip(),
                "weight": (export_info.get("weight") or "").strip(),
                "vehicle_trademark": (car365_info.get("vehicle_trademark") or "").strip(),
                "vehicle_first_registration": (car365_info.get("vehicle_first_registration") or "").strip(),
                "vehicle_fuel": (car365_info.get("vehicle_fuel") or "").strip(),
                "vehicle_displacement": (car365_info.get("vehicle_displacement") or "").strip(),
                "returned_chassis_no": (car365_info.get("returned_chassis_no") or "").strip(),
                "passport_status": passport_info.get("status", ""),
                "passport_error": passport_info.get("error", ""),
                "vin_status": vin_info.get("status", ""),
                "vin_error": vin_info.get("error", ""),
                "export_status": export_info.get("status", ""),
                "export_error": export_info.get("error", ""),
                "car365_status": car365_info.get("status", ""),
                "car365_error": car365_info.get("error", ""),
                "company_name": profile["company_name"] or "",
                "company_address": profile["company_address"] or "",
                "company_phone": profile["company_phone"] or "",
                "stamp_image_path": profile["stamp_image_path"] or "",
            }
        )

    output_filename = f"packing_list_{safe_filename_part(container_number)}_{request_id[:8]}.xlsx"
    output_path = OUTPUT_DIR / output_filename

    build_packing_list_workbook(output_path, vehicle_rows)

    return send_file(
        output_path,
        as_attachment=True,
        download_name=output_filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------- Excel ----------------

def build_packing_list_workbook(output_path: Path, rows: List[dict]) -> None:
    wb = load_workbook(TEMPLATE_PATH)
    template_sheet = wb[wb.sheetnames[0]]

    while len(wb.sheetnames) < len(rows):
        wb.copy_worksheet(template_sheet)

    while len(wb.sheetnames) > len(rows):
        wb.remove(wb[wb.sheetnames[-1]])

    for sheet_name, item in zip(wb.sheetnames, rows):
        ws = wb[sheet_name]
        fill_packing_sheet(ws, item)
        try:
            ws.title = item["kg_number"]
        except ValueError:
            pass

    wb.save(output_path)


def insert_stamp_image(ws, image_path: str):
    if not image_path:
        return
    path = Path(image_path)
    if not path.exists():
        return

    img = XLImage(str(path))
    img.width = 185
    img.height = 110
    ws.add_image(img, "E37")


def fill_packing_sheet(ws, item: dict) -> None:
    ws["B3"] = item.get("company_name", "")
    ws["B4"] = item.get("company_address", "")
    ws["B6"] = f"TEL: {item.get('company_phone', '')}"
    insert_stamp_image(ws, item.get("stamp_image_path", ""))

    ws["B8"] = f"Name: {item.get('full_name', '')}"
    ws["B9"] = f"Passport: {item.get('document_id', '')}"
    ws["B29"] = item.get("vin_string", "")
    ws["E16"] = f"CONTAINER No. {item.get('container_number', '')}"
    ws["E17"] = f"SEAL No. {item.get('seal_number', '')}"
    ws["G3"] = item.get("kg_number", "")
    ws["G4"] = item.get("page_date", "")
    ws["C17"] = item.get("destination", "")

    ws["A25"] = item.get("hs_code", "")

    weight_value = item.get("weight", "")
    if weight_value:
        try:
            ws["G25"] = float(weight_value)
        except ValueError:
            ws["G25"] = weight_value

    ws["B25"] = item.get("vehicle_trademark", "")
    ws["B26"] = item.get("vehicle_fuel", "")
    ws["B27"] = item.get("vehicle_first_registration", "")
    ws["B32"] = item.get("vehicle_displacement", "")

    if item.get("passport_status") == "ERROR" and item.get("passport_error"):
        ws["B12"] = f"Passport error: {item['passport_error']}"
    if item.get("vin_status") == "ERROR" and item.get("vin_error"):
        ws["B31"] = f"VIN error: {item['vin_error']}"
    if item.get("export_status") == "ERROR" and item.get("export_error"):
        ws["A27"] = f"Export error: {item['export_error']}"
    if item.get("car365_status") not in ("", "OK", "SKIPPED") and item.get("car365_error"):
        ws["B34"] = f"Car365 error: {item['car365_error']}"

@app.route("/analyze-vin", methods=["POST"])
def analyze_vin():

    try:
        files = request.files.getlist("files")

        if not files:
            return jsonify({
                "success": False,
                "error": "No files uploaded"
            })

        temp_dir = UPLOAD_DIR / uuid.uuid4().hex
        temp_dir.mkdir(parents=True, exist_ok=True)

        vin_result = ""

        for file in files:

            filename = secure_filename(file.filename)
            lower = filename.lower()

            path = temp_dir / filename
            file.save(path)

            # VIN 후보 파일만 처리
            if (
                "vin" in lower
                or "sticker" in lower
                or "label" in lower
            ):
                vin_info = extract_vin_info(path)
                vin_result = vin_info.get("vin_string", "")

        return jsonify({
            "success": True,
            "vin_string": vin_result
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })
    
    reader = easyocr.Reader(['en'], gpu=False)


@app.route("/analyze-passport", methods=["POST"])
def analyze_passport():

    try:
        files = request.files.getlist("files")

        if not files:
            return jsonify({"success": False, "error": "No files uploaded"})

        temp_dir = UPLOAD_DIR / uuid.uuid4().hex
        temp_dir.mkdir(parents=True, exist_ok=True)

        result_data = {
            "passport_no": "",
            "surname": "",
            "given_names": ""
        }

        for file in files:

            filename = secure_filename(file.filename)
            path = temp_dir / filename
            file.save(path)

            # OCR 실행
            ocr_results = reader.readtext(str(path), detail=0)

            clean_lines = [
                line.replace(" ", "").upper()
                for line in ocr_results
                if len(line.strip()) > 20
            ]

            if len(clean_lines) < 2:
                if len(ocr_results) >= 2:
                    clean_lines = [
                        line.replace(" ", "").upper()
                        for line in ocr_results[-2:]
                    ]
                else:
                    continue

            line1 = clean_lines[-2]
            line2 = clean_lines[-1]

            surname = "Unknown"
            given_names = "Unknown"

            if "<<" in line1:
                parts = line1.split("<<")
                left_part = parts[0]
                right_part = parts[1] if len(parts) > 1 else ""

                if "<" in left_part:
                    surname = left_part.split("<")[-1]

                given_names = right_part.replace("<", " ").strip()

            passport_no = line2[:9] if len(line2) >= 9 else line2

            # 가장 마지막 값 유지 (여러 파일 중)
            result_data = {
                "passport_no": passport_no,
                "surname": surname,
                "given_names": given_names
            }

        return jsonify({
            "success": True,
            **result_data
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })
    
if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)